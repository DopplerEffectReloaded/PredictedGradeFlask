from secret import db_details
import mysql.connector
import openpyxl

PE1_WEIGHTAGE = 0.1
PE2_WEIGHTAGE = 0.1
PE3_WEIGHTAGE = 0.3
PEY_WEIGHTAGE = 0.4

englishHL = [12, 25, 39, 53, 65, 79, 100]
englishSL = [10, 24, 37, 52, 64, 78, 100]
spanishBSL = [8, 19, 32, 48, 65, 80, 100]
frenchBSL = [9, 21, 35, 50, 66, 81, 100]
hindiBSL = [11, 25, 34, 50, 67, 83, 100]
spanishABSL = [11, 25, 39, 53, 66, 79, 100]
frenchABSL = [11, 25, 42, 54, 67, 79, 100]
math_aaHL = [12, 20, 29, 42, 56, 69, 100]
math_aiHL = [11, 22, 31, 43, 56, 67, 100]
math_aaSL = [8, 16, 28, 41, 56, 74, 100]
math_aiSL = [10, 21, 32, 47, 62, 76, 100]
physicsHL = [13, 23, 36, 46, 57, 68, 100]
physicsSL = [11, 21, 34, 44, 54, 64, 100]
chemistryHL = [15, 26, 38, 50, 63, 75, 100]
chemistrySL = [14, 27, 37, 48, 59, 71, 100]
biologyHL = [14, 24, 35, 49, 63, 76, 100]
biologySL = [14, 25, 37, 49, 61, 73, 100]
psychologyHL = [9, 21, 34, 48, 60, 73, 100]
psychologySL = [10, 23, 35, 47, 60, 73, 100]
digital_societiesHL = [12, 25, 40, 51, 60, 71, 100]
digital_societiesSL = [10, 22, 34, 46, 57, 69, 100]
global_politicsHL = [10, 22, 34, 46, 58, 70, 100]
global_politicsSL = [9, 21, 32, 42, 53, 63, 100]
visual_artsHL = [10, 21, 38, 52, 65, 80, 100]
visual_artsSL = [10, 21, 33, 50, 64, 80, 100]
economicsHL = [11, 24, 35, 47, 61, 73, 100]
economicsSL = [12, 25, 39, 50, 61, 72, 100]
business_managementHL = [15, 30, 40, 49, 58, 68, 100]
business_managementSL = [12, 25, 36, 48, 61, 72, 100]
computer_scienceHL = [14, 29, 40, 51, 61, 72, 100]
computer_scienceSL = [14, 29, 43, 53, 63, 73, 100]
essHL = [29, 39, 49, 59, 71, 84, 100]
essSL = [29, 39, 49, 59, 71, 84, 100]

subName = {
    'english' : 'English',
    'spanishB' : 'Spanish B',
    'frenchB' : 'French B',
    'hindiB' : 'Hindi B',
    'spanishAB' : 'Spanish AB',
    'frenchAB' : 'French AB',
    'math_aa' : 'Mathematics AA',
    'math_ai' : 'Mathematics AI',
    'physics' : 'Physics',
    'chemistry' : 'Chemistry',
    'biology' : 'Biology',
    'psychology' : 'Psychology',
    'digital_societies' : 'Digital Societies',
    'global_politics' : 'Global Politics',
    'visual_arts' : 'Visual Arts',
    'economics' : 'Economics',
    'business_management' : 'Business Management',
    'computer_science' : 'Computer Science',
    'ess' : 'ESS'
}

def Predictor(PE1, PE2, PE3, PE_Y2, grade_boundary, weightagePE1, weightagePE2, weightagePE3, weightagePE1Y2):
    
    if PE1 is None:
        return ('Incomplete information', 'Incomplete information', 'Incomplete information')
    
    if PE2 is None:

        current_grade = 1           
        for i in range(0, 7):
            
            d = grade_boundary[i]

            if weightagePE1*PE1 >= weightagePE1*d:
                current_grade += 1

        future_grades = []
        for i in range(current_grade, 8):
            boundary = grade_boundary[i - 2]
            marks = ((boundary*(weightagePE1+weightagePE2))-weightagePE1*PE1)/weightagePE2
            future_grades.append((i, round(marks, 1)))

        # future_grade  = current_grade + 1
        # if (future_grade - 7) >= 1:
        #     future_grade = 7

        # boundary = grade_boundary[future_grade-2]
        # future_marks = ((boundary*(weightagePE1+weightagePE2))-weightagePE1*PE1)/weightagePE2
        return ([current_grade, future_grades])

    elif PE3 is None:
        y = weightagePE1*PE1 + weightagePE2*PE2   
        current_grade = 1
        for i in range(0, 7):

            f = grade_boundary[i]

            if (weightagePE1*PE1 + weightagePE2*PE2) >= (weightagePE1*f) + (weightagePE2*f):
                current_grade += 1

        future_grades = []
        for i in range(current_grade, 8):
            boundary = grade_boundary[i - 2]
            marks = ((boundary*(weightagePE1+weightagePE2+weightagePE3)) - y)/weightagePE3
            future_grades.append((i, round(marks, 1)))
        
        # future_grade  = current_grade + 1
        # if (future_grade - 7) >= 1:
        #     future_grade = 7
        
        # boundary = grade_boundary[future_grade-2]
        # future_marks = ((boundary*(weightagePE1+weightagePE2+weightagePE3)) - y)/weightagePE3
        return ([current_grade, future_grades])
    
    elif PE_Y2 is None:
        z = weightagePE1*PE1 + weightagePE2*PE2 + weightagePE3*PE3
        current_grade = 1
        for i in range(0, 7):

            j = grade_boundary[i]

            if z >= (weightagePE1*j) + (weightagePE2*j) + (weightagePE3*j):
                current_grade += 1

        future_grades = []
        for i in range(current_grade, 8):
            boundary = grade_boundary[i - 2]
            marks = ((boundary*(weightagePE1+weightagePE2+weightagePE3+weightagePE1Y2)) - z)/weightagePE1Y2
            future_grades.append((i, round(marks, 1)))
        
        # future_grade  = current_grade + 1
        # if (future_grade - 7) >= 1:
        #     future_grade = 7
        
        # boundary = grade_boundary[future_grade-2]
        # future_marks = ((boundary*(weightagePE1+weightagePE2+weightagePE3+weightagePE1Y2)) - z)/weightagePE1Y2
        return ([current_grade, future_grades])
    else:
        z = weightagePE1*PE1 + weightagePE2*PE2 + weightagePE3*PE3 + weightagePE1Y2*PE_Y2
        current_grade = 1
        for i in range(0, 7):

            j = grade_boundary[i]

            if z >= (weightagePE1*j) + (weightagePE2*j) + (weightagePE3*j) + (weightagePE1Y2*j):
                current_grade += 1
        future_grade = 'N/A, all exams have been taken'
        future_marks = 'N/A, all exams have been taken'
        return ([current_grade, []])

def is_admin(email, password):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute(f'select password from admin where uid=%s', [email])
    table_password = cursor.fetchall()
    cursor.close()
    db.close()
    if len(table_password) == 0:
        return 2
    table_password = table_password[0][0]
    if password == table_password:
        return 1
    else:
        return 0

def is_student(email, password):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute(f'select password from student where uid=%s', [email])
    table_password = cursor.fetchall()
    cursor.close()
    db.close()
    if len(table_password) == 0:
        return False
    table_password = table_password[0][0]
    if password == table_password:
        return True
    else:
        return False

def list_to_str(array):
    string = str(array[0])
    for i in range(1, len(array)):
        string += f' {array[i]}'
    return string

def str_to_list(string):
    Array = string.split(' ')
    for i in range(len(Array)): Array[i] = int(Array[i]) 
    return Array

def set_boundary(subName, boundary_list):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    boundary_string = list_to_str(boundary_list)
    cursor.execute(f'delete from boundaries where subName=%s', [subName])
    cursor.execute(f'insert into boundaries (subName, boundary) values (%s, %s)', [subName, boundary_string])
    db.commit()
    cursor.close()
    db.close()

def get_boundary(subName):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute(f'select boundary from boundaries where subName=%s', [subName])
    boundary_string = cursor.fetchall()[0][0]
    cursor.close()
    db.close()
    boundary = str_to_list(boundary_string)
    return boundary

def parse_db(uid):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute(f'select * from marks where uid =%s', [uid])
    db_data = cursor.fetchall()[0][1:]
    cursor.close()
    db.close()
    
    temp_1 = Predictor(db_data[1], db_data[2], db_data[3], db_data[4], get_boundary(db_data[0]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
    temp_2 = Predictor(db_data[6], db_data[7], db_data[8], db_data[9], get_boundary(db_data[5]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
    temp_3 = Predictor(db_data[11], db_data[12], db_data[13], db_data[14], get_boundary(db_data[10]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
    temp_4 = Predictor(db_data[16], db_data[17], db_data[18], db_data[19], get_boundary(db_data[15]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
    temp_5 = Predictor(db_data[21], db_data[22], db_data[23], db_data[24], get_boundary(db_data[20]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
    temp_6 = Predictor(db_data[26], db_data[27], db_data[28], db_data[29], get_boundary(db_data[25]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)

    form_values = {     'subject_1': f'{subName[db_data[0][:-2]]} {db_data[0][-2:]}',
                        'sub1_PE1': db_data[1],
                        'sub1_PE2': db_data[2],
                        'sub1_PE3': db_data[3],
                        'sub1_PEY2': db_data[4],
                        'sub1_grade': temp_1[0],
                        'sub1_next_grades': temp_1[1],
                        'subject_2': f'{subName[db_data[5][:-2]]} {db_data[5][-2:]}',
                        'sub2_PE1': db_data[6],
                        'sub2_PE2': db_data[7],
                        'sub2_PE3': db_data[8],
                        'sub2_PEY2': db_data[9],
                        'sub2_grade': temp_2[0],
                        'sub2_next_grades': temp_2[1],
                        'subject_3': f'{subName[db_data[10][:-2]]} {db_data[10][-2:]}',
                        'sub3_PE1': db_data[11],
                        'sub3_PE2': db_data[12],
                        'sub3_PE3': db_data[13],
                        'sub3_PEY2': db_data[14],
                        'sub3_grade': temp_3[0],
                        'sub3_next_grades': temp_3[1],
                        'subject_4': f'{subName[db_data[15][:-2]]} {db_data[15][-2:]}',
                        'sub4_PE1': db_data[16],
                        'sub4_PE2': db_data[17],
                        'sub4_PE3': db_data[18],
                        'sub4_PEY2': db_data[19],
                        'sub4_grade': temp_4[0],
                        'sub4_next_grades': temp_4[1],
                        'subject_5': f'{subName[db_data[20][:-2]]} {db_data[20][-2:]}',
                        'sub5_PE1': db_data[21],
                        'sub5_PE2': db_data[22],
                        'sub5_PE3': db_data[23],
                        'sub5_PEY2': db_data[24],
                        'sub5_grade': temp_5[0],
                        'sub5_next_grades': temp_5[1],
                        'subject_6': f'{subName[db_data[25][:-2]]} {db_data[25][-2:]}',
                        'sub6_PE1': db_data[26],
                        'sub6_PE2': db_data[27],
                        'sub6_PE3': db_data[28],
                        'sub6_PEY2': db_data[29],
                        'sub6_grade': temp_6[0],
                        'sub6_next_grades': temp_6[1]
                        }
    return form_values

def get_grades(subject):
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    grades = []
    cursor.execute(f'select uid, sub1_pe1, sub1_pe2, sub1_pe3, sub1_pey2 from marks where sub1 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.execute(f'select uid, sub2_pe1, sub2_pe2, sub2_pe3, sub2_pey2 from marks where sub2 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.execute(f'select uid, sub3_pe1, sub3_pe2, sub3_pe3, sub3_pey2 from marks where sub3 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.execute(f'select uid, sub4_pe1, sub4_pe2, sub4_pe3, sub4_pey2 from marks where sub4 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.execute(f'select uid, sub5_pe1, sub5_pe2, sub5_pe3, sub5_pey2 from marks where sub5 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.execute(f'select uid, sub6_pe1, sub6_pe2, sub6_pe3, sub6_pey2 from marks where sub6 = %s', [subject])
    for i in cursor.fetchall(): grades.append(i)
    cursor.close()
    db.close()
    for i in range(len(grades)):
        grades[i] = list(grades[i])
        grades[i].append(Predictor(grades[i][1], grades[i][2], grades[i][3], grades[i][4], get_boundary(subject), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)[0])
    return grades

def get_final_grades():
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute(f'select *from marks')
    db_data = []
    for i in cursor.fetchall():
        name_and_grades = []
        grades = []
        name_and_grades.append(i[0])
        
        temp1 = Predictor(i[2], i[3], i[4], i[5], get_boundary(i[1]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp2 = Predictor(i[7], i[8], i[9], i[10], get_boundary(i[6]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp3 = Predictor(i[12], i[13], i[14], i[15], get_boundary(i[11]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp4 = Predictor(i[17], i[18], i[19], i[20], get_boundary(i[16]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp5 = Predictor(i[22], i[23], i[24], i[25], get_boundary(i[21]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp6 = Predictor(i[27], i[28], i[29], i[30], get_boundary(i[26]), PE1_WEIGHTAGE, PE2_WEIGHTAGE, PE3_WEIGHTAGE, PEY_WEIGHTAGE)
        temp1.pop()
        temp2.pop()
        temp3.pop()
        temp4.pop()
        temp5.pop()
        temp6.pop()
        grades.append(temp1[0])
        grades.append(temp2[0])
        grades.append(temp3[0])
        grades.append(temp4[0])
        grades.append(temp5[0])
        grades.append(temp6[0])
        try:
            predictedgrade = grades[0] + grades[1] + grades[2] + grades[3] + grades[4] + grades[5]
            name_and_grades.append(predictedgrade)
            db_data.append(name_and_grades)
        except TypeError:
            continue
    cursor.close()
    db.close()
    return db_data

subNameExcel = {
    'english': 'english',
    'spanish b': 'spanishB',
    'french b': 'frenchB',
    'hindi b': 'hindiB',
    'spanish ab': 'spanishAB',
    'french ab': 'frenchAB',
    'math aa': 'math_aa',
    'math ai': 'math_ai',
    'physics': 'physics',
    'chemistry': 'chemistry',
    'biology': 'biology',
    'psychology': 'psychology',
    'digital societies': 'digital_societies',
    'global politics': 'global_politics',
    'visual arts': 'visual_arts',
    'economics': 'economics',
    'business management': 'business_management',
    'bm': 'business_management',
    'computer science': 'computer_science',
    'cs': 'computer_science',
    'ess': 'ess'
}

def get_grade(min_row, data):
    arr = [[] for _ in range(6)]
    if min_row == 0:
        arr.append(data.cell(2, 2).value)
        for i in range(6):
            arr[i].append(data.cell(2, i+3).value)
    else:
        arr.append(data.cell(min_row, 2).value)
        for i in range(6):
            arr[i].append(data.cell(min_row, i+3).value)

    check = ['Progression 1 Marks out of 100', 'Progression 2 Marks Out of 100', 'Progression 3 Marks out of 100', 'Progression 1 of year 2 out of 100']
    
    index = 0
    for i in data.iter_rows(min_row, data.max_row):
        if i[1].value in check:
            for subj in range(2, 8):
                arr[subj-2].append(i[subj].value)    
            index += 1

        if i[1].value is None:
            break
        
    return arr

def student_grade(data):
    full_arr = []
    full_arr.append(get_grade(0, data))

    flag = 0
    min_row = 1
    
    for i in data.iter_rows(0, data.max_row):
        min_row +=1 

        if i[1].value is None:
            flag += 1
            full_arr.append(get_grade(min_row, data))
        else:
            flag = 0

        if flag == 2:
            break
    
    empty_row = [[None], [None], [None], [None], [None], [None], None]
    while True:
        if empty_row in full_arr:
            full_arr.remove(empty_row)
        else:
            break

    parameters= []
    for i in full_arr:

        parameters.append((i[-1].strip() if i[-1].strip() else 'NULL', 
                           subNameExcel[i[0][0].strip().lower()]+'HL', i[0][1] if i[0][1] is not None else 'NULL', i[0][2] if i[0][2] is not None else 'NULL', i[0][3] if i[0][3] is not None else 'NULL', i[0][4] if i[0][4] is not None else 'NULL',
                           subNameExcel[i[1][0].strip().lower()]+'HL', i[1][1] if i[1][1] is not None else 'NULL', i[1][2] if i[1][2] is not None else 'NULL', i[1][3] if i[1][3] is not None else 'NULL', i[1][4] if i[1][4] is not None else 'NULL',
                           subNameExcel[i[2][0].strip().lower()]+'HL', i[2][1] if i[2][1] is not None else 'NULL', i[2][2] if i[2][2] is not None else 'NULL', i[2][3] if i[2][3] is not None else 'NULL', i[2][4] if i[2][4] is not None else 'NULL',
                           subNameExcel[i[3][0].strip().lower()]+'SL', i[3][1] if i[3][1] is not None else 'NULL', i[3][2] if i[3][2] is not None else 'NULL', i[3][3] if i[3][3] is not None else 'NULL', i[3][4] if i[3][4] is not None else 'NULL',
                           subNameExcel[i[4][0].strip().lower()]+'SL', i[4][1] if i[4][1] is not None else 'NULL', i[4][2] if i[4][2] is not None else 'NULL', i[4][3] if i[4][3] is not None else 'NULL', i[4][4] if i[4][4] is not None else 'NULL',
                           subNameExcel[i[5][0].strip().lower()]+'SL', i[5][1] if i[5][1] is not None else 'NULL', i[5][2] if i[5][2] is not None else 'NULL', i[5][3] if i[5][3] is not None else 'NULL', i[5][4] if i[5][4] is not None else 'NULL',
                           ))
        
    db = mysql.connector.connect(**db_details)
    cursor = db.cursor()
    cursor.execute('delete from marks')
    db.commit()
    for i in parameters:
        query = "insert into marks (uid, sub1, sub1_pe1, sub1_pe2, sub1_pe3, sub1_pey2, sub2, sub2_pe1, sub2_pe2, sub2_pe3, sub2_pey2, sub3, sub3_pe1, sub3_pe2, sub3_pe3, sub3_pey2, sub4, sub4_pe1, sub4_pe2, sub4_pe3, sub4_pey2, sub5, sub5_pe1, sub5_pe2, sub5_pe3, sub5_pey2, sub6, sub6_pe1, sub6_pe2, sub6_pe3, sub6_pey2) values ('%s', '%s', %s, %s, %s, %s,'%s', %s, %s, %s, %s, '%s', %s, %s, %s, %s, '%s', %s, %s, %s, %s, '%s', %s, %s, %s, %s, '%s', %s, %s, %s, %s)"
        cursor.execute(query % i)
    
    db.commit()
    cursor.close()
    db.close()
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                               
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                  
                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                  
